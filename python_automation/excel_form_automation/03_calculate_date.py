from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook


class WeeklyWorkPlan:
    wb = None
    ws = None
    start_date = '2025-12-08'
    manager = "매니저 이름을 입력 해주세요"
    date_list = []
    days_of_week = []

    def __init__(self, start_date, manager, sheet_no=0):
        self.start_date = start_date
        self.manager = manager
        self.wb = Workbook()
        self.ws = self.wb.active
        self.set_title()

        # 날짜 생성
        self.set_date()


    def save(self, filename):
        self.wb.save(filename)
        print('엑셀파일 생성 완료')

    def set_date(self, days=6):
        # start_date + 6일
        end_date = datetime.strptime(self.start_date, '%Y-%m-%d') + timedelta(days=days)

        # [5 6 7 8 9 10 11]
        week = pd.date_range(start=self.start_date, end=end_date.strftime('%Y-%m-%d'))
        self.date_list = week.strftime('%Y-%m-%d').to_list()
        self.days_of_week = week.strftime('%A').to_list()

        print('end_date:', end_date)
        print('week:', week)
        print('date_list:', self.date_list)

    def set_title(self):
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = '2025-12-08'
        end_date = '2025-12-14'
        ws['B6'] = f'({start_date}-{end_date})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')
        print('타이틀 생성 완료')




if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2025-12-08','김성수')

    wwp.save('result/주간업무계획표.xlsx')